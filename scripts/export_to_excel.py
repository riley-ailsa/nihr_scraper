#!/usr/bin/env python3
"""
Scrape NIHR funding opportunities and export to Excel.

Usage:
    python scrape_to_excel.py [--limit N] [--output FILENAME]

Examples:
    python scrape_to_excel.py                    # Scrape with defaults
    python scrape_to_excel.py --limit 5          # Scrape 5 opportunities
    python scrape_to_excel.py --output test.xlsx # Custom output file
"""

import sys
import argparse
import logging
from pathlib import Path
from datetime import datetime

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from src.ingest.nihr_funding import NihrFundingScraper
from src.normalize.nihr import normalize_nihr_opportunity, infer_nihr_status

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def load_urls_from_file(filepath: str = "nihr_urls.txt") -> list:
    """
    Load opportunity URLs from a text file.

    Args:
        filepath: Path to the URLs file

    Returns:
        List of URLs (ignores comments and blank lines)
    """
    urls = []
    file_path = Path(__file__).parent / filepath

    if not file_path.exists():
        logger.warning(f"URL file not found: {file_path}")
        return urls

    with open(file_path, 'r') as f:
        for line in f:
            line = line.strip()
            # Skip empty lines and comments
            if line and not line.startswith('#'):
                urls.append(line)

    return urls


def scrape_opportunities(urls: list, limit: int = None) -> list:
    """
    Scrape opportunity data from NIHR.

    Args:
        urls: List of opportunity URLs to scrape
        limit: Maximum number to scrape (None = all)

    Returns:
        List of scraped opportunity data dicts
    """
    scraper = NihrFundingScraper()
    results = []

    urls_to_process = urls[:limit] if limit else urls

    for i, url in enumerate(urls_to_process, 1):
        logger.info(f"[{i}/{len(urls_to_process)}] Scraping: {url}")

        try:
            # Scrape the opportunity
            opportunity = scraper.scrape(url)

            # Normalize to get computed fields
            grant, indexable_docs = normalize_nihr_opportunity(opportunity)

            # Get status
            status = infer_nihr_status(opportunity)

            # Convert to dict for Excel export
            result = {
                'id': opportunity.opportunity_id,
                'reference_id': opportunity.reference_id or '',
                'title': grant.title,
                'programme': opportunity.programme or '',
                'type': opportunity.opportunity_type or '',
                'status': status,
                'description': (grant.description[:500] + '...') if grant.description and len(grant.description) > 500 else (grant.description or ''),
                'url': grant.url,
                'opens_at': opportunity.opening_date.strftime('%Y-%m-%d %H:%M') if opportunity.opening_date else '',
                'closes_at': opportunity.closing_date.strftime('%Y-%m-%d %H:%M') if opportunity.closing_date else '',
                'funding': opportunity.funding_text or '',
                'funding_gbp': grant.total_fund_gbp or '',
                'sections_count': len(opportunity.sections),
                'resources_count': len(opportunity.resources),
                'tags': ', '.join(grant.tags) if grant.tags else '',
            }

            results.append(result)
            logger.info(f"  OK: {grant.title[:60]}...")

        except Exception as e:
            logger.error(f"  ERROR: {e}")
            continue

    return results


def export_to_excel(data: list, filename: str):
    """
    Export scraped data to Excel file.

    Args:
        data: List of opportunity dicts
        filename: Output Excel filename
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "NIHR Opportunities"

    # Define columns
    columns = [
        ('ID', 20),
        ('Reference', 12),
        ('Title', 50),
        ('Programme', 30),
        ('Type', 15),
        ('Status', 10),
        ('Description', 60),
        ('URL', 50),
        ('Opens', 18),
        ('Closes', 18),
        ('Funding', 30),
        ('Funding (GBP)', 15),
        ('Sections', 10),
        ('Resources', 10),
        ('Tags', 30),
    ]

    # Header row styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    # Write headers
    for col_idx, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    for row_idx, record in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=record.get('id', ''))
        ws.cell(row=row_idx, column=2, value=record.get('reference_id', ''))
        ws.cell(row=row_idx, column=3, value=record.get('title', ''))
        ws.cell(row=row_idx, column=4, value=record.get('programme', ''))
        ws.cell(row=row_idx, column=5, value=record.get('type', ''))
        ws.cell(row=row_idx, column=6, value=record.get('status', ''))
        ws.cell(row=row_idx, column=7, value=record.get('description', ''))
        ws.cell(row=row_idx, column=8, value=record.get('url', ''))
        ws.cell(row=row_idx, column=9, value=record.get('opens_at', ''))
        ws.cell(row=row_idx, column=10, value=record.get('closes_at', ''))
        ws.cell(row=row_idx, column=11, value=record.get('funding', ''))
        ws.cell(row=row_idx, column=12, value=record.get('funding_gbp', ''))
        ws.cell(row=row_idx, column=13, value=record.get('sections_count', 0))
        ws.cell(row=row_idx, column=14, value=record.get('resources_count', 0))
        ws.cell(row=row_idx, column=15, value=record.get('tags', ''))

        # Wrap text for description
        ws.cell(row=row_idx, column=7).alignment = Alignment(wrap_text=True, vertical='top')

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save
    wb.save(filename)
    logger.info(f"Excel file saved: {filename}")


def main():
    parser = argparse.ArgumentParser(description='Scrape NIHR opportunities to Excel')
    parser.add_argument('--limit', type=int, default=None,
                        help='Maximum number of opportunities to scrape')
    parser.add_argument('--output', type=str, default=None,
                        help='Output Excel filename (default: auto-generated)')
    parser.add_argument('--urls-file', type=str, default='nihr_urls.txt',
                        help='Path to URLs file (default: nihr_urls.txt)')
    args = parser.parse_args()

    print("=" * 60)
    print("NIHR SCRAPER - Excel Export")
    print("=" * 60)
    print()

    # Load URLs from file
    urls = load_urls_from_file(args.urls_file)

    if not urls:
        logger.error(f"No opportunity URLs found. Add URLs to {args.urls_file}")
        sys.exit(1)

    logger.info(f"Loaded {len(urls)} URLs from {args.urls_file}")

    # Scrape opportunities
    logger.info(f"Scraping {args.limit or len(urls)} opportunities...")
    results = scrape_opportunities(urls, limit=args.limit)

    if not results:
        logger.error("No opportunities scraped successfully!")
        sys.exit(1)

    # Generate output filename
    if args.output:
        filename = args.output
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"nihr_opportunities_{timestamp}.xlsx"

    # Export to Excel
    export_to_excel(results, filename)

    # Summary
    print()
    print("=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Opportunities scraped: {len(results)}")
    print(f"Output file: {filename}")
    print()


if __name__ == "__main__":
    main()
